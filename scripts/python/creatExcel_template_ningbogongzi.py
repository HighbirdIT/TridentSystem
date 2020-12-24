import json
import sys
import io
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font
import json
import sys
import os
import shutil


sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
argv = sys.argv
#argv = ['',r'd:\work\TridentSystem\filedata\excel\097d5152-5fd8-552f-8e90-b94f612a6f07.xlsx',r'd:\work\TridentSystem\filedata\exceljson\097d5152-5fd8-552f-8e90-b94f612a6f07.json',r'd:\work\TridentSystem\filedata\exceltemplate\baoshuigongzi.xlsx']
filePath = str(argv[1])
jsonPath = str(argv[2])
templetePath = str(argv[3])

with open(jsonPath, 'r', encoding='utf-8') as f:
    json_data=json.loads(f.read())

shutil.copyfile(templetePath,filePath)

wb=load_workbook(filePath)
sheet = wb.active

rows = json_data['rows']
for rowIndex in range(0, len(rows)):
        rowdata = rows[rowIndex]
        useRowIndex=rowIndex+5
        if rowIndex == 0:
            工资月份 = int(rowdata['工资月份'])
            cell = sheet.cell(row=2, column=2, value=(rowdata['工资年份'] + (str(工资月份) if 工资月份 >= 10 else '0' + str(工资月份))))
        cell = sheet.cell(row=useRowIndex, column=1, value=(rowdata['人员姓名'] if '人员姓名' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=2, value=(rowdata['银行账号'] if '银行账号' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=3, value=(rowdata['金额'] if '金额' in rowdata else ''))
        

wb.save(filePath)
print('OK',end='')


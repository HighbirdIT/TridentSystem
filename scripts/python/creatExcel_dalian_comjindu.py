from sqlAccess import SqlData
from style import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import typing
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.pagebreak import Break
import openpyxl
import os
import math
import time
import json
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
argv = sys.argv
# argv = ['',r'd:\work\TridentSystem\filedata\excel\F2构件进度追踪1112.xlsx',r'D:\work\python\铝条表格\test.json']
excelPath = str(argv[1])
jsonPath = str(argv[2])

def getColumnName(index):
    key = ''
    if index > 26:
        modVal = index % 26
        if modVal == 0:
            index -= 1
            modVal = 26
        key = chr(64 + math.floor(index / 26))
        index = modVal
    index -= 1
    return key + chr(65 + index)
def GenBorder(left_style, right_style, top_style, bottom_style):
    rltBorder = Border()
    if left_style != None:
        rltBorder.left = Side(border_style=left_style,color='000000')
    if right_style != None:
        rltBorder.right = Side(border_style=right_style,color='000000')
    if right_style != None:
        rltBorder.top = Side(border_style=top_style,color='000000')
    if bottom_style != None:
        rltBorder.bottom = Side(border_style=bottom_style,color='000000')
    return rltBorder

thinSide = Side(border_style='thin',color='000000')
mediumSide = Side(border_style='medium',color='000000')
dashedSide = Side(border_style='dashed',color='000000')
doubleSide = Side(border_style='double',color='000000')
mediumDashedSide = Side(border_style='mediumDashed',color='000000')

titleBorder = GenBorder('thin','thin','medium','medium')
titleBorder_left = GenBorder('thin','dashed','medium','medium')
titleBorder_right = GenBorder('dashed','thin','medium','medium')
titleBorder_mid = GenBorder('dashed','dashed','medium','medium')

cellBorder = GenBorder('thin','thin','thin','thin')

headerBorder = Border(left=Side(border_style='medium',color='000000'),
                    right=Side(border_style='medium',color='000000'),
                    top=Side(border_style='medium',color='000000'),
                    bottom=Side(border_style='medium',color='000000'))


with open(jsonPath, 'r', encoding='utf-8') as f:
    json_data=json.loads(f.read())

Row_Start = json_data['行Range'][0]
Row_End = json_data['行Range'][1]
Row_Count = Row_End - Row_Start + 1
Col_Start = json_data['列Range'][0]
Col_End = json_data['列Range'][1]
Col_Count = Col_End - Col_Start + 1
level = json_data['level']
queryDate = json_data['queryDate']

headerFontStyle = Font(size = "20")
infoFontStyle = Font(size = "15")

cellAlign = Alignment(vertical="center",horizontal='center',wrapText=False)

grayFill = PatternFill(fill_type = 'solid',start_color='919191',end_color='919191')
redFill = PatternFill(fill_type = 'solid',start_color='FF0000',end_color='FF0000')
greenFill = PatternFill(fill_type = 'solid',start_color='00FF00',end_color='00FF00')
blueFill = PatternFill(fill_type = 'solid',start_color='22A7FF',end_color='22A7FF')
whiteFill = PatternFill(fill_type = 'solid',start_color='FFFFFF',end_color='FFFFFF')

def write工序(targetSheet,baseRow,cell_dic,title,工序):
    row_index = baseRow
    targetSheet.merge_cells('A%d:%s%d'%(row_index,getColumnName(Col_Count+1),row_index))
    headerCell = targetSheet.cell(row=row_index, column=1, value="%s的%s进度    查询时间:%s"%(title,工序,queryDate))
    headerCell.font = headerFontStyle
    for i in range(1,Col_Count + 2):
        targetSheet['%s%d'%(getColumnName(i),row_index)].border = headerBorder
    
    
    row_index += 1
    infoRowIndex = row_index
    targetSheet.merge_cells('A%d:%s%d'%(row_index,getColumnName(Col_Count+1),row_index))

    # wirteCol
    row_index += 1
    targetSheet.row_dimensions[row_index].height = 18
    for i in range(0, Col_Count + 1):
        if i == 0:
            colLabel = ''
        else:
            colLabel = Col_Start + (i - 1)
        cell = targetSheet.cell(row=row_index, column=i+1,value=colLabel)
        cell.border = titleBorder
        cell.alignment = cellAlign
        if i == 0:
            cell.fill = grayFill
    
    count = 0
    doneCount = 0
    todyCount = 0
    yesterDayCount = 0
    for row_i in range(1,Row_Count+1):
        row_index += 1
        targetSheet.row_dimensions[row_index].height = 18
        realRow = Row_End - (row_i - 1)
        rowLabel = realRow

        cell = targetSheet.cell(row=row_index, column=1,value=rowLabel)
        cell.border = titleBorder
        cell.alignment = cellAlign

        for col_i in range(1, Col_Count + 1):
            realCol = Col_Start + (col_i - 1)
            cellKey = '%s_%s'%(realRow,realCol)
            cellData = None
            colLabel = ''
            useFill = grayFill
            if cellKey in cell_dic:
                count += 1
                cellData = cell_dic[cellKey]
                工序Data = cellData[工序]
                if 工序Data['isDone']:
                    doneCount += 1
                    useFill = greenFill
                    passDay = 工序Data['passDay']
                    if passDay > 9:
                        colLabel = '9+'
                    else:
                        colLabel = passDay
                        if passDay == 0:
                            todyCount += 1
                            useFill = blueFill
                            colLabel = ''
                        elif passDay == 1:
                            yesterDayCount += 1
                else:
                    useFill = redFill

            cell = targetSheet.cell(row=row_index, column=col_i+1,value=colLabel)
            cell.border = cellBorder
            cell.alignment = cellAlign
            cell.fill = useFill
            
    infoText = '%s进度:%d/%d  今日完成:%d  昨日完成:%d'%(工序,doneCount,count,todyCount,yesterDayCount)
    headerCell = targetSheet.cell(row=infoRowIndex, column=1, value=infoText)
    headerCell.font = infoFontStyle
    for i in range(1,Col_Count + 2):
        targetSheet['%s%d'%(getColumnName(i),row_index)].border = cellBorder
    return row_index


wb =  Workbook()
del wb._sheets[0]

def wirt构件(name):
    sheet = wb.create_sheet('%s%s'%(level,name))
    for i in range(1,Col_Count + 2):
        sheet.column_dimensions[getColumnName(i)].width = 4
    构件data = json_data[name]
    工序_arr = 构件data['工序_arr']
    baseRow = 1
    for 工序 in 工序_arr:
        baseRow = write工序(sheet,baseRow,构件data,'%s%s'%(level,name),工序)
        baseRow += 2

wirt构件('钢框架')
wirt构件('铝框架')
wirt构件('驳接爪')
wirt构件('特殊驳接爪')

wb.save(excelPath)

print('OK',end='')

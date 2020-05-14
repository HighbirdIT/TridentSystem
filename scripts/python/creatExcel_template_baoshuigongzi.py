import json
import sys
import io
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font
import json
import sys
import os
import shutil


sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
argv = sys.argv
#argv = ['',r'd:\work\TridentSystem\filedata\excel\097d5152-5fd8-552f-8e90-b94f612a6f07.xlsx',r'd:\work\TridentSystem\filedata\exceljson\097d5152-5fd8-552f-8e90-b94f612a6f07.json',r'd:\work\TridentSystem\filedata\exceltemplate\baoshuigongzi.xlsx']
filePath = str(argv[1])
jsonPath = str(argv[2])
templetePath = str(argv[3])

with open(jsonPath, 'r', encoding='utf-8') as f:
    json_data=json.loads(f.read())

shutil.copyfile(templetePath,filePath)

wb=load_workbook(filePath)
sheet = wb.active

rows = json_data['rows']
for rowIndex in range(0, len(rows)):
        rowdata = rows[rowIndex]
        useRowIndex=rowIndex+2
        cell = sheet.cell(row=useRowIndex, column=1, value=(rowdata['工号'] if '工号' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=2, value=(rowdata['姓名'] if '姓名' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=3, value=(rowdata['证照类型'] if '证照类型' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=4, value=(rowdata['证照号码'] if '证照号码' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=5, value=(rowdata['收入额'] if '收入额' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=6, value=(''))
        cell = sheet.cell(row=useRowIndex, column=7, value=(rowdata['养老'] if '养老' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=8, value=(rowdata['医疗'] if '医疗' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=9, value=(rowdata['失业'] if '失业' in rowdata else ''))
        cell = sheet.cell(row=useRowIndex, column=10, value=(rowdata['公积金'] if '公积金' in rowdata else ''))
        

wb.save(filePath)
print('OK',end='')


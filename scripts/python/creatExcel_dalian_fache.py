from sqlAccess import SqlData
from style import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import typing
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill
from openpyxl.worksheet.pagebreak import Break
import openpyxl
import datetime
import math
import time
import json
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
argv = sys.argv
# argv = ['',r'd:\work\TridentSystem\filedata\excel\F2构件进度追踪1112.xlsx',r'D:\work\python\铝条表格\test.json']
excelPath = str(argv[1])
jsonPath = str(argv[2])


def getColumnName(index):
    key = ''
    if index > 26:
        modVal = index % 26
        if modVal == 0:
            index -= 1
            modVal = 26
        key = chr(64 + math.floor(index / 26))
        index = modVal
    index -= 1
    return key + chr(65 + index)


def GenBorder(left_style, right_style, top_style, bottom_style):
    rltBorder = Border()
    if left_style != None:
        rltBorder.left = Side(border_style=left_style, color='000000')
    if right_style != None:
        rltBorder.right = Side(border_style=right_style, color='000000')
    if right_style != None:
        rltBorder.top = Side(border_style=top_style, color='000000')
    if bottom_style != None:
        rltBorder.bottom = Side(border_style=bottom_style, color='000000')
    return rltBorder


thinSide = Side(border_style='thin', color='000000')
mediumSide = Side(border_style='medium', color='000000')
dashedSide = Side(border_style='dashed', color='000000')
doubleSide = Side(border_style='double', color='000000')
mediumDashedSide = Side(border_style='mediumDashed', color='000000')

titleBorder = GenBorder('thin', 'thin', 'medium', 'medium')
titleBorder_left = GenBorder('thin', 'dashed', 'medium', 'medium')
titleBorder_right = GenBorder('dashed', 'thin', 'medium', 'medium')
titleBorder_mid = GenBorder('dashed', 'dashed', 'medium', 'medium')

cellBorder = GenBorder('thin', 'thin', 'thin', 'thin')

headerBorder = Border(left=Side(border_style='medium', color='000000'),
                      right=Side(border_style='medium', color='000000'),
                      top=Side(border_style='medium', color='000000'),
                      bottom=Side(border_style='medium', color='000000'))
header1FontStyle = Font(size="26", bold=True)
header2FontStyle = Font(size="14", bold=True)
header3FontStyle = Font(size="12", bold=True)
infoFontStyle = Font(size="15")
infoBoldFontStyle = Font(size="11", bold=True)
cellAlign = Alignment(vertical="center", horizontal='center', wrapText=False)
cellRightAlign = Alignment(
    vertical="center", horizontal='right', wrapText=False)
cellAlignwrap = Alignment(
    vertical="center", horizontal='center', wrapText=True)
grayFill = PatternFill(
    fill_type='solid', start_color='919191', end_color='919191')
redFill = PatternFill(
    fill_type='solid', start_color='FF0000', end_color='FF0000')
greenFill = PatternFill(
    fill_type='solid', start_color='00FF00', end_color='00FF00')
blueFill = PatternFill(
    fill_type='solid', start_color='22A7FF', end_color='22A7FF')
whiteFill = PatternFill(
    fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')


def write标题头(targetSheet, 发货标题, 发货日期, 标题头行数, 列数):
    for i in range(1, 标题头行数 + 1):
        targetSheet.merge_cells('A%d:%s%d' % (i, getColumnName(列数), i))
        tableCell = targetSheet.cell(row=i, column=1)
        tableCell.alignment = cellAlign
        if i==1:
            tableCell.value = "上海海勃膜结构股份有限公司"
            tableCell.font = header1FontStyle
        elif i==2:
            tableCell.value = 发货标题
            tableCell.font = header2FontStyle
        elif i==3:
            tableCell.value = ("发货日期:%s") % (发货日期)
            tableCell.font = header3FontStyle
            tableCell.alignment = cellRightAlign


def write构件列标题(targetSheet, 构件列标题, 当前行数):
    row位置 = 当前行数 + 1
    for i in range(1, 7):
        j = i-4 if i >= 4 else i-1
        TitleCell = targetSheet.cell(row=row位置, column=i)
        TitleCell.value = 构件列标题[j]
        TitleCell.font = infoBoldFontStyle
        TitleCell.alignment = cellAlign
        targetSheet.row_dimensions[row位置].height = 20
        targetSheet['%s%d' % (getColumnName(i), row位置)].border = cellBorder


def write车辆信息(targetSheet, data, 标题头行数, 货车信息行数, 列数):
    """ 发货记录 = 发货信息(data) """
    发货对象 = [["货车公司:", data[0]['货车公司名称'], "车牌号码:", data[0]['车牌号码']],
            ["驾驶人:", data[0]['驾驶人'], "手机号:", data[0]['手机号码']],
            ["包总数量:", data[0]['包数量'], "构件总数量:", data[0]['构件数量']],
            ["备注说明:", data[0]['备注说明'], '', '']]
    for i in range(0, 货车信息行数):
        row位置 = i + 标题头行数 + 1
        targetSheet.row_dimensions[row位置].height = 20
        for k in range(1, 列数 + 1):
            targetSheet['%s%d' % (getColumnName(k), row位置)].border = cellBorder
        if len(发货对象) > i:
            col位置arr = [1, 3, 5, 6]
            targetSheet.merge_cells('%s%d:%s%d' % (
                getColumnName(1), row位置, getColumnName(2), row位置))
            if i >= 3:
                col位置arr = col位置arr[0:2]
                targetSheet.merge_cells('%s%d:%s%d' % (
                    getColumnName(3), row位置, getColumnName(6), row位置))
            else:
                targetSheet.merge_cells('%s%d:%s%d' % (
                    getColumnName(3), row位置, getColumnName(4), row位置))
            for col位置 in col位置arr:
                rowCell = targetSheet.cell(row=row位置, column=col位置)
                rowCell.font = infoBoldFontStyle
                rowCell.alignment = cellAlign
                if col位置 == 1:
                    rowCell.value = 发货对象[i][0]
                if col位置 == 3:
                    rowCell.value = 发货对象[i][1]
                if col位置 == 5:
                    rowCell.value = 发货对象[i][2]
                if col位置 == 6:
                    rowCell.value = 发货对象[i][3]


包分界序号 = 0
def 货物打包列表(targetSheet, 发货记录代码, 当前行数):
    sql = ("select  distinct 货物打包记录代码,货物打包名称,登记确认时间  from FT258C大连发货明细(%s)  order by 登记确认时间" % (发货记录代码))
    sql_res = SqlData(sql=sql)
    global rowNum1, rowNum2, 序号
    rowNum1 = 当前行数
    rowNum2 = 当前行数
    序号 = 0
    包分界序号 = len(sql_res.data) // 2
    for dataIndex, dataRow in enumerate(sql_res.data):
        货物打包记录代码 = dataRow['货物打包记录代码']
        write构件列表(targetSheet, 发货记录代码, 货物打包记录代码, 当前行数,dataIndex)
def write构件列表(targetSheet, 发货记录代码, 打包记录代码, 当前行数, 包序号):
    sql = ("""select   货物打包名称,构件记录代码,构件全称缓存,全局代码,部位代码,方位名称,一级序号,二级序号  from FT258C大连发货明细(%s) 
                where 货物打包记录代码 = %s
                order by 全局代码,部位代码,方位名称,一级序号,二级序号""" % (发货记录代码, 打包记录代码))
    sql_res = SqlData(sql=sql)
    num = 0
    global rowNum1, rowNum2, 序号

    ms_arr = []
    startRow = -1
    preCol = -1
    preRow = -1
    for dataIndex, dataRow in enumerate(sql_res.data):
        num = num + 1
        序号 = 序号 + 1
        #构件名称 = [序号, dataRow['构件全称缓存']]
        打包名称 = dataRow['货物打包名称']
        if 包序号 % 2 == 1:
        # if (int(序号/27)) % 2 == 1:
            rowNum2 = rowNum2 + 1
            rowNum = rowNum2
            col2 = 4
        else:
            rowNum1 = rowNum1 + 1
            rowNum = rowNum1
            col2 = 1
        if startRow == -1:
            startRow = rowNum
        if preCol != -1 and preCol != col2:
            ms_arr.append((preCol,startRow,preRow,"包:%s\n数量:%s" %(打包名称,preRow - startRow + 1)))
            startRow = rowNum
        targetSheet.row_dimensions[rowNum].height = 20
        cell = targetSheet.cell(row=rowNum, column=col2, value=序号)
        cell.border = cellBorder
        cell.alignment = cellAlign
        cell = targetSheet.cell(row=rowNum, column=col2+1, value=dataRow['构件全称缓存'])
        cell.border = cellBorder
        cell.alignment = cellAlign
        targetSheet.cell(row=rowNum, column=col2+2).border = cellBorder
        preRow = rowNum
        preCol = col2
    ms_arr.append((preCol,startRow,preRow,"包:%s\n数量:%s" %(打包名称,preRow - startRow + 1)))
    for setting in ms_arr:
        colName = getColumnName(setting[0] + 2)
        targetSheet.merge_cells('%s%d:%s%d'%(colName,setting[1],colName,setting[2]))
        打包方式Cell = targetSheet.cell(row=setting[1],column=setting[0] + 2,value=setting[3])
        打包方式Cell.alignment = cellAlignwrap
        打包方式Cell.font = infoBoldFontStyle

with open(jsonPath, 'r', encoding='utf-8') as f:
    json_data = json.loads(f.read())
发货记录代码 = json_data['发货记录代码']
sql = ("""select distinct t1.大连发货记录代码,货车公司名称,车牌号码,驾驶人,手机号码,发货日期,备注说明,包数量,构件数量 from FT258C大连发货明细(%s) as t1 inner join (select  发货记录代码,count(*)as 包数量 from T258C货物打包记录 group by 发货记录代码)as t2  on 大连发货记录代码 = t2.发货记录代码 inner join (select  大连发货记录代码,count(*)as 构件数量 from FT258C大连发货明细(%s) group by 大连发货记录代码)as t3  on t1.大连发货记录代码 = t3.大连发货记录代码""" 
          % (发货记录代码,发货记录代码))
sql_res = SqlData(sql=sql)
if sql_res.length > 0:
    wb = Workbook()
    sheet = wb.active
    标题头行数 = 3
    货车信息行数 = 4
    列数 = 6
    构件列标题 = ["序号", "构件名称", "打包方式"]
    for i in range(1, 列数+1):
        if i == 1 or i == 4:
            sheet.column_dimensions[getColumnName(i)].width = 3.7
            continue
        sheet.column_dimensions[getColumnName(i)].width = 20
    write标题头(sheet, "大连梭鱼湾足球场发货清单", sql_res.data[0]["发货日期"], 标题头行数, 列数)
    write车辆信息(sheet, sql_res.data, 标题头行数, 货车信息行数, 列数)
    write构件列标题(sheet, 构件列标题, 标题头行数+货车信息行数)
    货物打包列表(sheet, 发货记录代码, 8)
    wb.save(excelPath)
print('OK',end='')
import json
import sys
import io
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font
import json
import sys
import os
import shutil


sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
argv = sys.argv
#argv = ['',r'd:\work\TridentSystem\filedata\excel\b520c44a-c5a0-e0c0-48de-4139b08aa93d.xlsx',r'd:\work\TridentSystem\filedata\exceljson\f2c82ec3-ac95-f2e3-9f04-ae8ae70b78be.json',r'd:\work\TridentSystem\filedata\exceltemplate\huamingceA.xlsx']
filePath = str(argv[1])
jsonPath = str(argv[2])
templetePath = str(argv[3])

with open(jsonPath, 'r', encoding='utf-8') as f:
    json_data=json.loads(f.read())

shutil.copyfile(templetePath,filePath)

wb=load_workbook(filePath)
sheet = wb.active

rows = json_data['rows']
上海rowIndex = 3
江苏rowindex = 3
for rowIndex in range(0, len(rows)):
        rowdata = rows[rowIndex]
        公司名称 = rowdata['公司名称'] if '公司名称' in rowdata else ''
        useRowIndex = -1
        useColIndex = -1
        if 公司名称 == '上海海勃':
            useRowIndex = 上海rowIndex
            useColIndex = 1
            上海rowIndex+=1
        elif 公司名称 == '江苏海勃':
            useRowIndex = 江苏rowindex
            useColIndex = 11
            江苏rowindex+=1
        if useRowIndex > 0:
            cell = sheet.cell(row=useRowIndex, column=useColIndex, value=(useRowIndex - 2) )
            cell = sheet.cell(row=useRowIndex, column=useColIndex + 1, value=(rowdata['员工姓名'] if '员工姓名' in rowdata else ''))
            cell = sheet.cell(row=useRowIndex, column=useColIndex + 2, value=(rowdata['工时状态'] if '工时状态' in rowdata else ''))
            cell = sheet.cell(row=useRowIndex, column=useColIndex + 3, value=(rowdata['系统名称'] if '系统名称' in rowdata else ''))
            cell = sheet.cell(row=useRowIndex, column=useColIndex + 4, value=(rowdata['部门名称'] if '部门名称' in rowdata else ''))
        

wb.save(filePath)
print('OK',end='')

